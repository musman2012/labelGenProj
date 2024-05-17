import os
from flask import render_template, request, send_file
from app import app
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            pdf_path = generate_pdf_from_excel(filepath)
            return send_file(pdf_path, as_attachment=True)
    return render_template('upload.html')

def generate_pdf_from_excel(file_path):
    # Same function you have for reading Excel and generating PDF
    articles = read_excel(file_path)
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'articles_and_prices.pdf')
    pdf_path = 'C:\\Users\\muhammad usman\\code\\labelGen\\labelGenProj\\uploads\\articles_and_prices.pdf'
    generate_pdf(articles, pdf_path)
    return pdf_path

def read_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    articles = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title = row[3]
        price = row[11]
        articles.append((title, price))
    return articles

def generate_pdf(articles, output_path):
    currency = "€"
    c = canvas.Canvas(output_path, pagesize=letter)
    page_width, page_height = letter
    pdfmetrics.registerFont(TTFont('Times-Roman', 'Times-RomanRegular.ttf'))
    pdfmetrics.registerFont(TTFont('Times-Bold', 'titles.ttf'))
    title_font = 'Times-Roman'
    price_font = 'Times-Bold'
    font_size = 20
    title_font_size = 15
    small_title_size = 13

    x_start = 0.5 * inch  # Starting x position
    y_start = page_height - 0.5 * inch  # Starting y position (from top)
    min_box_width = 4 * cm  # Minimum box width to accommodate title and price
    padding = 1 * cm  # Padding from left and right of the box
    item_height = 1.5 * inch

    x = x_start
    y = y_start

    for i, (title, price) in enumerate(articles):
        if title == 'New' or price == 0:
          continue

        text_width = c.stringWidth(title, title_font, title_font_size)
        box_width = max(min_box_width, text_width + 2 * padding)
        if x + box_width > page_width - 0.5 * inch:
            x = x_start
            y -= item_height
        if y < 0.5 * inch + item_height:
            c.showPage()
            x = x_start
            y = y_start
        
        # Draw the boundary box
        c.rect(x, y - item_height, box_width, item_height)

        # Draw the title
        c.setFont(title_font, title_font_size)
        c.drawString(x + padding, y - 0.5 * inch, title)

        # Draw the price centered
        c.setFont(price_font, font_size)
        formatted_price = f"{price:.2f} {currency}"
        price_width = c.stringWidth(formatted_price, price_font, font_size)
        c.drawString(x + (box_width - price_width) / 2, y - 0.9 * inch, formatted_price)

        x += box_width
    c.save()

